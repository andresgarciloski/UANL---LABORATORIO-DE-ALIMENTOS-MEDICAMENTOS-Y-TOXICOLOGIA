import os
import tempfile
import datetime
from tkinter import messagebox, filedialog
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from core.auth import agregar_historial
import pandas as pd

def escribir_celda_segura(ws, coord, valor):
    """Escribe en una celda asegurando que, si est√° dentro de un rango merged, se escriba en la celda superior izquierda."""
    if not isinstance(coord, str):
        try:
            coord = coord.coordinate
        except Exception:
            coord = str(coord)
    for rango in ws.merged_cells.ranges:
        if coord in rango:
            ws.cell(row=rango.min_row, column=rango.min_col).value = valor
            return
    ws[coord] = valor

def _sanitize_filename(name: str) -> str:
    """Quitar caracteres inv√°lidos y espacios duplicados para filenames."""
    keep = "-_.() abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
    cleaned = "".join(c for c in name if c in keep)
    cleaned = "_".join(part for part in cleaned.split() if part)
    return cleaned.strip("_") or "archivo"

class NutrimentalExporter:
    def __init__(self, parent_window):
        self.parent = parent_window

    def agregar_sellos_advertencia(self, ws, resultados):
        """Inserta im√°genes de sellos en celdas espec√≠ficas del template siguiendo info2.txt.
        Las im√°genes se colocan con tama√±o mayor al de las celdas pero no modifican
        el alto/ancho de filas/columnas del libro (se superponen)."""
        # Obtener flags desde la UI si est√°n disponibles
        es_liquida = False
        es_bebida_sin_calorias = False
        try:
            tipo_var = getattr(self.parent, "tipo_muestra", None)
            if tipo_var is not None and tipo_var.get() == "liquida":
                es_liquida = True
        except Exception:
            es_liquida = False
        try:
            beb_var = getattr(self.parent, "bebida_sin_calorias", None)
            if beb_var is not None and beb_var.get():
                es_bebida_sin_calorias = True
        except Exception:
            es_bebida_sin_calorias = False

        por100 = resultados.get("por_100g", {})

        # Valores esperados
        calorias = float(por100.get("energia_kcal", 0) or 0)
        azucares_g = float(por100.get("azucares", 0) or 0)
        grasa_sat_g = float(por100.get("grasa_saturada", 0) or 0)
        grasa_trans_mg = float(por100.get("grasa_trans", 0) or 0)
        sodio_mg = float(por100.get("sodio", 0) or 0)

        # Calcular sellos seg√∫n info2.txt
        sellos = {
            "exceso_calorias": False,
            "exceso_azucares": False,
            "exceso_grasas_saturadas": False,
            "exceso_grasas_trans": False,
            "exceso_sodio": False
        }

        # EXCESO CALOR√çAS
        if es_liquida:
            if calorias >= 70:
                sellos["exceso_calorias"] = True
        else:
            if calorias >= 275:
                sellos["exceso_calorias"] = True
        if (azucares_g * 4) >= 8:
            sellos["exceso_calorias"] = True

        # EXCESO AZ√öCARES (azucares_g * 4 >= 10% de kcal)
        if calorias > 0 and (azucares_g * 4) >= (0.10 * calorias):
            sellos["exceso_azucares"] = True

        # EXCESO GRASAS SATURADAS (grasa_sat_g * 9 >= 10% de kcal)
        if calorias > 0 and (grasa_sat_g * 9) >= (0.10 * calorias):
            sellos["exceso_grasas_saturadas"] = True

        # EXCESO GRASAS TRANS (grasa_trans_mg /1000 *9 >= 1% de kcal)
        if calorias > 0:
            energia_trans_kcal = (grasa_trans_mg / 1000.0) * 9.0
            if energia_trans_kcal >= (0.01 * calorias):
                sellos["exceso_grasas_trans"] = True

        # EXCESO SODIO (>=300 mg OR mg sodio >= kcal OR bebida sin calor√≠as >=45 mg)
        if sodio_mg >= 300:
            sellos["exceso_sodio"] = True
        elif calorias >= 0 and sodio_mg >= calorias:
            sellos["exceso_sodio"] = True
        elif es_bebida_sin_calorias and sodio_mg >= 45:
            sellos["exceso_sodio"] = True

        # Mapeo de sello -> imagen, celda y tama√±o deseado (px)
        project_root = os.path.dirname(os.path.dirname(__file__))
        ruta_base = os.path.abspath(os.path.join(project_root, "img", "Sellos"))
        sellos_config = {
            "exceso_azucares": {"imagen": "azucares.jpg", "celda": "M3", "size": (160,160)},
            "exceso_calorias": {"imagen": "calorias.jpg", "celda": "O3", "size": (160,160)},
            "exceso_grasas_saturadas": {"imagen": "saturadas.jpg", "celda": "Q3", "size": (160,160)},
            # SODIO y GRASAS TRANS ahora en fila 14 (M14 y O14)
            "exceso_sodio": {"imagen": "sodio.jpg", "celda": "M14", "size": (140,140)},
            "exceso_grasas_trans": {"imagen": "trans.jpg", "celda": "O14", "size": (140,140)}
        }

        for key, aplica in sellos.items():
            if not aplica:
                continue
            cfg = sellos_config.get(key)
            if not cfg:
                continue
            ruta_imagen = os.path.join(ruta_base, cfg["imagen"])
            if not os.path.exists(ruta_imagen):
                print(f"[exporter] imagen no encontrada: {ruta_imagen}")
                continue
            try:
                img = XLImage(ruta_imagen)
                # tama√±o mayor sin tocar celdas: establece width/height m√°s grandes
                width_px, height_px = cfg.get("size", (140,140))
                img.width = width_px
                img.height = height_px
                # agregar imagen anclada en la celda solicitada; no modificar√° filas/columnas
                ws.add_image(img, cfg["celda"])
            except Exception as e:
                print(f"[exporter] fallo al insertar imagen {ruta_imagen}: {e}")

    def llenar_plantilla_excel(self, wb, resultados, entrada, datos_basicos):
        ws = wb.active

        # Determinar unidad autom√°tica: "mL" si es l√≠quida, "g" si es s√≥lida (fallback a "g")
        unidad = "g"
        tipo_var = getattr(self.parent, "tipo_muestra", None)
        try:
            if tipo_var is not None and tipo_var.get() == "liquida":
                unidad = "mL"
        except Exception:
            unidad = "g"

        # Tama√±o de porci√≥n con unidad en F17 (escribe s√≥lo la celda destino)
        porcion_val = entrada.get("porcion", "")
        escribir_celda_segura(ws, "F17", f"{porcion_val} {unidad}" if porcion_val != "" else "")

        # Porciones por envase en F18 (escribe s√≥lo la celda destino), sin limpiar columna C
        porciones_envase = resultados.get("porciones_envase", None)
        if porciones_envase is not None and porciones_envase != "":
            try:
                pv = float(porciones_envase)
                porciones_display = int(pv) if pv.is_integer() else round(pv, 1)
            except Exception:
                porciones_display = porciones_envase
            escribir_celda_segura(ws, "F18", f"{porciones_display} {unidad}")
        else:
            escribir_celda_segura(ws, "F18", "")

        # NO limpiar celdas de la columna C ni otras etiquetas de la plantilla
        # (anteriormente se estaban borrando C17/C18 y eso eliminaba etiquetas)

        # Contenido energ√©tico por envase y contenido neto con unidad (F19 y F12)
        escribir_celda_segura(ws, "F19", f"{resultados.get('por_envase', {}).get('energia_kcal','')} kcal")
        contenido_neto = entrada.get("contenido_neto", "")
        escribir_celda_segura(ws, "F12", f"{contenido_neto} {unidad}" if contenido_neto != "" else "")

        # Nombre y descripci√≥n
        escribir_celda_segura(ws, "C8", f"{datos_basicos.get('nombre','')} - {datos_basicos.get('descripcion','')}")

        # Mapeo valores por 100g y por porci√≥n (mantener resto igual)
        m = resultados.get("por_100g", {})
        p = resultados.get("por_porcion", {})
        mappings = [
            ("F21", m.get("energia_kcal", ""), "G21", "kcal"),
            ("F22", m.get("proteina", ""), "G22", "g"),
            ("F23", m.get("grasa_total", ""), "G23", "g"),
            ("F24", m.get("grasa_saturada", ""), "G24", "g"),
            ("F25", m.get("grasa_trans", ""), "G25", "mg"),
            ("F26", m.get("carbohidratos_disponibles", ""), "G26", "g"),
            ("F27", m.get("azucares", ""), "G27", "g"),
            ("F28", m.get("azucares_anadidos", ""), "G28", "g"),
            ("F29", m.get("fibra_dietetica", ""), "G29", "g"),
            ("F30", m.get("sodio", ""), "G30", "mg"),
            ("H21", p.get("energia_kcal", ""), "I21", "kcal"),
            ("H22", p.get("proteina", ""), "I22", "g"),
            ("H23", p.get("grasa_total", ""), "I23", "g"),
            ("H24", p.get("grasa_saturada", ""), "I24", "g"),
            ("H25", p.get("grasa_trans", ""), "I25", "mg"),
            ("H26", p.get("carbohidratos_disponibles", ""), "I26", "g"),
            ("H27", p.get("azucares", ""), "I27", "g"),
            ("H28", p.get("azucares_anadidos", ""), "I28", "g"),
            ("H29", p.get("fibra_dietetica", ""), "I29", "g"),
            ("H30", p.get("sodio", ""), "I30", "mg"),
        ]
        for cel_val_100, val100, cel_unit, unit in mappings:
            try:
                escribir_celda_segura(ws, cel_val_100, val100)
            except Exception:
                ws[cel_val_100] = val100
            try:
                ws[cel_unit] = unit
            except Exception:
                pass

        # Agregar sellos de advertencia (usa c√°lculo en nutrimental.py)
        self.agregar_sellos_advertencia(ws, resultados)
        return wb

    def generar_pdf_desde_excel(self, wb, nombre_archivo, guardar_dialogo=True):
        """Genera PDF usando Excel COM; incluye initialfile en el di√°logo de guardado."""
        tmp_excel_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                wb.save(tmp.name)
                tmp_excel_path = tmp.name

            try:
                import win32com.client
            except ImportError:
                raise ImportError("win32com no disponible. Instale pywin32 para exportar a PDF desde Excel en Windows.")

            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            wb_x = excel.Workbooks.Open(tmp_excel_path)

            if guardar_dialogo:
                # incluir nombre sugerido en initialfile
                ruta_guardado_pdf = filedialog.asksaveasfilename(
                    defaultextension=".pdf",
                    initialfile=nombre_archivo,
                    filetypes=[("PDF","*.pdf")]
                )
                if not ruta_guardado_pdf:
                    wb_x.Close(False)
                    excel.Quit()
                    return None, "Cancelado"
            else:
                ruta_guardado_pdf = os.path.join(tempfile.gettempdir(), nombre_archivo)

            wb_x.ExportAsFixedFormat(0, ruta_guardado_pdf)
            wb_x.Close(False)
            excel.Quit()

            with open(ruta_guardado_pdf, "rb") as f:
                contenido = f.read()
            return contenido, ruta_guardado_pdf

        finally:
            if tmp_excel_path and os.path.exists(tmp_excel_path):
                try:
                    os.remove(tmp_excel_path)
                except Exception:
                    pass

    def exportar_a_formato_predefinido(self):
        """Exporta usando plantilla y sugiere un nombre estandarizado seg√∫n datos ingresados."""
        if not hasattr(self.parent, "ultimo_calculo"):
            messagebox.showwarning("Advertencia", "Primero debe calcular la tabla nutrimental")
            return
        try:
            plantilla = "formato.xlsx"
            if not os.path.exists(plantilla):
                raise FileNotFoundError("Plantilla formato.xlsx no encontrada.")
            wb = load_workbook(plantilla)
            resultados = self.parent.ultimo_calculo["resultados"]
            entrada = self.parent.ultimo_calculo["datos_entrada"]
            datos_basicos = self.parent.ultimo_calculo["datos_basicos"]

            # Generar nombre recomendado estandarizado
            nombre_raw = datos_basicos.get("nombre", "muestra")
            nombre_limpio = _sanitize_filename(str(nombre_raw))
            tipo = "liquida" if getattr(self.parent, "tipo_muestra", None) and self.parent.tipo_muestra.get() == "liquida" else "solida"
            # unidad y porcion (sin decimales si es entero)
            unidad = "mL" if tipo == "liquida" else "g"
            porcion_val = entrada.get("porcion", "")
            try:
                pv = float(porcion_val)
                porcion_str = str(int(pv)) if pv.is_integer() else str(pv).replace(".", "_")
            except Exception:
                porcion_str = _sanitize_filename(str(porcion_val))
            fecha = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            recommended = f"{nombre_limpio}_{tipo}_{porcion_str}{unidad}_{fecha}.pdf"
            recommended = _sanitize_filename(recommended)

            wb = self.llenar_plantilla_excel(wb, resultados, entrada, datos_basicos)

            # pasar el nombre recomendado a la funci√≥n que muestra el di√°logo
            _, ruta = self.generar_pdf_desde_excel(wb, recommended, True)
            if ruta != "Cancelado":
                messagebox.showinfo("√âxito", f"Archivo PDF exportado correctamente:\n\n{ruta}")
        except FileNotFoundError:
            messagebox.showerror("Error", "No se encontr√≥ la plantilla formato.xlsx.")
        except ImportError as ie:
            messagebox.showwarning("PDF no generado", str(ie))
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo completar la exportaci√≥n:\n{e}")

    def guardar_solo_bd(self):
        if not hasattr(self.parent, "ultimo_calculo"):
            messagebox.showwarning("Advertencia", "Primero debe calcular la tabla nutrimental")
            return
        try:
            nombre_actual = self.parent.nombre_entry.get().strip()
            descripcion_actual = self.parent.descripcion_entry.get("1.0", "end-1c").strip()
            if not nombre_actual:
                messagebox.showerror("Error", "El campo '# de muestra' es obligatorio para guardar")
                self.parent.nombre_entry.focus()
                return
            usuario_id = self.parent.get_usuario_id()
            if usuario_id is None:
                messagebox.showwarning("Advertencia", "No se pudo guardar en la base de datos: Usuario no v√°lido")
                return
            plantilla = "formato.xlsx"
            if not os.path.exists(plantilla):
                raise FileNotFoundError("Plantilla formato.xlsx no encontrada.")
            wb = load_workbook(plantilla)
            resultados = self.parent.ultimo_calculo["resultados"]
            entrada = self.parent.ultimo_calculo["datos_entrada"]
            datos_basicos = self.parent.ultimo_calculo["datos_basicos"]
            wb = self.llenar_plantilla_excel(wb, resultados, entrada, datos_basicos)
            fecha = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_pdf = f"nutrimental_{nombre_actual}_{fecha}.pdf"
            archivo_binario, _ = self.generar_pdf_desde_excel(wb, nombre_pdf, False)
            if not archivo_binario:
                messagebox.showwarning("Advertencia", "No se pudo generar el PDF para la base de datos.")
                return
            fecha_actual = datetime.datetime.now()
            descripcion_completa = f"{descripcion_actual}\n\nTABLA NUTRIMENTAL OFICIAL (PDF):\n"
            descripcion_completa += f"- Prote√≠na: {resultados['por_100g'].get('proteina','')}g/100g\n"
            descripcion_completa += f"- Grasa total: {resultados['por_100g'].get('grasa_total','')}g/100g\n"
            descripcion_completa += f"- Carbohidratos disponibles: {resultados['por_100g'].get('carbohidratos_disponibles','')}g/100g\n"
            descripcion_completa += f"- Energ√≠a: {resultados['por_100g'].get('energia_kcal','')} kcal/100g\n"
            descripcion_completa += f"- Tama√±o de porci√≥n: {entrada.get('porcion','')}g\n"
            descripcion_completa += f"- Contenido neto: {entrada.get('contenido_neto','No especificado')}\n"
            descripcion_completa += f"Archivo PDF formato oficial generado el {fecha_actual.strftime('%d/%m/%Y %H:%M:%S')}"
            agregar_historial(
                nombre_actual,
                descripcion_completa,
                fecha_actual.strftime("%Y-%m-%d"),
                fecha_actual.strftime("%H:%M:%S"),
                usuario_id,
                archivo_binario
            )
            messagebox.showinfo("Guardado en Base de Datos", "‚úÖ PDF nutrimental guardado correctamente en la base de datos.")
        except FileNotFoundError:
            messagebox.showerror("Error", "No se encontr√≥ la plantilla formato.xlsx.")
        except ImportError as ie:
            messagebox.showwarning("PDF no generado", str(ie))
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo completar la exportaci√≥n:\n{e}")

    def exportar_nutrimental_excel(self):
        if not hasattr(self.parent, "ultimo_calculo"):
            messagebox.showwarning("Advertencia", "Primero debe calcular la tabla nutrimental")
            return
        try:
            nombre_actual = self.parent.nombre_entry.get().strip()
            descripcion_actual = self.parent.descripcion_entry.get("1.0", "end-1c").strip()
            if not nombre_actual:
                messagebox.showerror("Error", "El campo '# de muestra' es obligatorio para exportar")
                self.parent.nombre_entry.focus()
                return
            fecha_actual = datetime.datetime.now()
            nombre_limpio = "".join(c for c in nombre_actual if c.isalnum() or c in (' ', '-', '_')).rstrip().replace(' ','_') or "Tabla_Nutrimental"
            timestamp = fecha_actual.strftime("%Y-%m-%d_%H%M%S")
            nombre_predeterminado = f"{nombre_limpio}_{timestamp}.xlsx"
            datos_excel = []
            datos_excel.append(["INFORMACI√ìN B√ÅSICA","",""])
            datos_excel.append(["# de muestra", nombre_actual, ""])
            datos_excel.append(["Descripci√≥n", descripcion_actual, ""])
            datos_excel.append(["Fecha de an√°lisis", fecha_actual.strftime("%Y-%m-%d"), ""])
            datos_excel.append(["Hora de an√°lisis", fecha_actual.strftime("%H:%M:%S"), ""])
            datos_excel.append(["Usuario", getattr(self.parent, "username", ""), ""])
            datos_excel.append(["Fecha de exportaci√≥n", fecha_actual.strftime("%Y-%m-%d %H:%M:%S"), ""])
            datos_excel.append(["","",""])
            datos_excel.append(["DATOS DE ENTRADA","",""])
            for key, value in self.parent.ultimo_calculo["datos_entrada"].items():
                nombre_campo = key.replace('_',' ').title()
                if key in ("sodio","grasa_trans"):
                    datos_excel.append([f"{nombre_campo} (mg/100g)", value, ""])
                elif key == "porcion":
                    datos_excel.append([f"{nombre_campo} (g)", value, ""])
                elif key == "contenido_neto":
                    datos_excel.append([f"{nombre_campo} (g/mL)", value, ""])
                else:
                    datos_excel.append([f"{nombre_campo} (%)", value, ""])
            datos_excel.append(["","",""])
            datos_excel.append(["TABLA NUTRIMENTAL MEXICANA","Por 100g/mL","Por Porci√≥n"])
            resultados = self.parent.ultimo_calculo["resultados"]
            if "porciones_envase" in resultados and resultados["porciones_envase"] is not None:
                datos_excel.append(["Porciones por envase", resultados["porciones_envase"], ""])
            for key in resultados["por_100g"].keys():
                if key == "energia_kcal": nombre = "Contenido energ√©tico (kcal)"
                elif key == "energia_kj": nombre = "Contenido energ√©tico (kJ)"
                elif key == "sodio": nombre = "Sodio (mg)"
                elif key == "grasa_trans": nombre = "Grasas trans (mg)"
                elif key == "azucares_anadidos": nombre = "Az√∫cares a√±adidos (g)"
                else: nombre = f"{key.replace('_',' ').title()} (g)"
                valor_100g = resultados["por_100g"][key]
                valor_porcion = resultados["por_porcion"][key]
                datos_excel.append([nombre, valor_100g, valor_porcion])
            if "por_envase" in resultados:
                datos_excel.append(["","",""]); datos_excel.append(["POR ENVASE COMPLETO","",""])
                for key, value in resultados["por_envase"].items():
                    if key == "energia_kcal": nombre = "Contenido energ√©tico total (kcal)"
                    elif key == "energia_kj": nombre = "Contenido energ√©tico total (kJ)"
                    else: nombre = f"{key.replace('_',' ').title()}"
                    datos_excel.append([nombre, value, ""])
            datos_excel.append(["","",""]); datos_excel.append(["SELLOS DE ADVERTENCIA","",""])
            sellos = {}
            calc = getattr(self.parent, "_calcular_sellos_advertencia", None)
            if callable(calc):
                sellos = calc(resultados)
            for sello, aplica in sellos.items():
                nombre_sello = sello.replace("exceso_", "EXCESO DE ").upper()
                datos_excel.append([nombre_sello, "S√ç" if aplica else "NO", ""])
            df = pd.DataFrame(datos_excel, columns=["Componente","Valor 100g/mL","Valor Porci√≥n"])
            try:
                desktop = os.path.join(os.path.expanduser("~"), "Desktop")
                if not os.path.exists(desktop):
                    desktop = os.path.expanduser("~")
            except:
                desktop = ""
            filename = filedialog.asksaveasfilename(initialfile=nombre_predeterminado, initialdir=desktop, defaultextension=".xlsx", filetypes=[("Excel files","*.xlsx"),("All files","*.*")], title="Guardar tabla nutrimental")
            if filename:
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name="Tabla Nutrimental")
                    workbook = writer.book
                    worksheet = writer.sheets["Tabla Nutrimental"]
                    worksheet.column_dimensions['A'].width = 30
                    worksheet.column_dimensions['B'].width = 15
                    worksheet.column_dimensions['C'].width = 15
                with open(filename, "rb") as f:
                    archivo_binario = f.read()
                usuario_id = self.parent.get_usuario_id()
                if usuario_id is None:
                    messagebox.showwarning("Advertencia", "No se pudo guardar en la base de datos: Usuario no v√°lido")
                    return
                descripcion_completa = f"{descripcion_actual}\n\nDATOS NUTRICIONALES CALCULADOS:\n"
                descripcion_completa += f"- Prote√≠na: {resultados['por_100g'].get('proteina','')}g/100g\n"
                descripcion_completa += f"- Grasa total: {resultados['por_100g'].get('grasa_total','')}g/100g\n"
                descripcion_completa += f"- Carbohidratos disponibles: {resultados['por_100g'].get('carbohidratos_disponibles','')}g/100g\n"
                descripcion_completa += f"- Energ√≠a: {resultados['por_100g'].get('energia_kcal','')} kcal/100g\n"
                descripcion_completa += f"- Tama√±o de porci√≥n analizada: {self.parent.ultimo_calculo['datos_entrada']['porcion']}g\n"
                descripcion_completa += f"Archivo Excel generado autom√°ticamente: {os.path.basename(filename)}"
                agregar_historial(
                    nombre_actual,
                    descripcion_completa,
                    fecha_actual.strftime("%Y-%m-%d"),
                    fecha_actual.strftime("%H:%M:%S"),
                    usuario_id,
                    archivo_binario
                )
                messagebox.showinfo("Exportaci√≥n Exitosa", f"‚úÖ Tabla nutrimental exportada correctamente:\n\nüìÅ Archivo: {os.path.basename(filename)}\nüìÇ Ubicaci√≥n: {filename}\nüíæ Guardado en base de datos: ‚úì")
            else:
                messagebox.showinfo("Cancelado", "Exportaci√≥n cancelada por el usuario.")
        except ImportError:
            messagebox.showerror("Error", "No se pudo importar pandas. Aseg√∫rate de que est√© instalado:\npip install pandas openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {e}")