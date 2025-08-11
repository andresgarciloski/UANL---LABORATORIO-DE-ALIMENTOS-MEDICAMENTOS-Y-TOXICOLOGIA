import tkinter as tk
from tkinter import messagebox, filedialog
import datetime
import os
import pandas as pd
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageTk
from core.auth import agregar_historial
from ui.base_interface import bind_mousewheel
import tempfile

def escribir_celda_segura(ws, celda, valor):
    """Escribe en una celda de forma segura incluso si es parte de una combinada"""
    # Obtener coordenadas de la celda
    if isinstance(celda, str):
        coord = celda
    else:
        coord = celda.coordinate
    
    # Comprobar si la celda está en un rango combinado
    for rango in ws.merged_cells.ranges:
        if coord in rango:
            # Si está en un rango combinado, usar la celda principal (esquina superior izquierda)
            celda_principal = ws.cell(row=rango.min_row, column=rango.min_col)
            celda_principal.value = valor
            return
    
    # Si no es una celda combinada, escribir normalmente
    ws[coord] = valor

class NutrimentalModule:
    def __init__(self, parent_window):
        self.parent = parent_window

    def show_nutrimental_section(self):
        """Mostrar sección de tabla nutrimental con diseño moderno y profesional mejorado"""
        for widget in self.parent.content_frame.winfo_children():
            widget.destroy()

        main_frame = tk.Frame(self.parent.content_frame, bg="#f4f8fc", bd=0)
        main_frame.pack(expand=True, fill="both", padx=20, pady=(5, 20))

        canvas = tk.Canvas(main_frame, bg="#f4f8fc", highlightthickness=0)
        scrollbar = tk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        scrollable_frame = tk.Frame(canvas, bg="#f4f8fc")
        window_id = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        def resize_inner_frame(event):
            canvas.itemconfig(window_id, width=event.width)
            canvas.configure(scrollregion=canvas.bbox("all"))
        canvas.bind("<Configure>", resize_inner_frame)

        # --- MEJOR SCROLL CON MOUSE SOLO EN EL CANVAS ---
        def _on_mousewheel(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except tk.TclError:
                pass  # El canvas ya no existe

        def _on_mousewheel_linux_up(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(-1, "units")  # Scroll hacia arriba
            except tk.TclError:
                pass  # El canvas ya no existe

        def _on_mousewheel_linux_down(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(1, "units")  # Scroll hacia abajo
            except tk.TclError:
                pass  # El canvas ya no existe

        # Vincular solo cuando el mouse está sobre el canvas
        canvas.bind("<Enter>", lambda e: canvas.bind("<MouseWheel>", _on_mousewheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind("<MouseWheel>"))
        canvas.bind("<Enter>", lambda e: canvas.bind("<Button-4>", _on_mousewheel_linux_up))
        canvas.bind("<Leave>", lambda e: canvas.unbind("<Button-4>"))
        canvas.bind("<Enter>", lambda e: canvas.bind("<Button-5>", _on_mousewheel_linux_down))
        canvas.bind("<Leave>", lambda e: canvas.unbind("<Button-5>"))

        # --- FIN MEJOR SCROLL ---

        # Configuración de columnas para que se adapten al ancho
        scrollable_frame.grid_columnconfigure(0, weight=1, minsize=220)
        scrollable_frame.grid_columnconfigure(1, weight=1, minsize=280)
        scrollable_frame.grid_columnconfigure(2, weight=2, minsize=300)
        scrollable_frame.grid_rowconfigure(0, weight=1)

        # Usar grid para los tres recuadros principales
        left_col = tk.Frame(scrollable_frame, bg="#f4f8fc")
        center_col = tk.Frame(scrollable_frame, bg="#f4f8fc")
        right_col = tk.Frame(scrollable_frame, bg="#f4f8fc")

        left_col.grid(row=0, column=0, sticky="nsew", padx=(5, 5), pady=5)
        center_col.grid(row=0, column=1, sticky="nsew", padx=(5, 5), pady=5)
        right_col.grid(row=0, column=2, sticky="nsew", padx=(5, 5), pady=5)

        self._create_basic_fields(left_col)
        self._create_nutrimental_fields(center_col)
        self._create_results_area(right_col)

        # Botones debajo de los recuadros, ocupando todo el ancho
        buttons_frame = tk.Frame(scrollable_frame, bg="#f4f8fc")
        buttons_frame.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(10, 10))
        
        # Centrar botones
        buttons_container = tk.Frame(buttons_frame, bg="#f4f8fc")
        buttons_container.pack(expand=True)
        self._create_buttons(buttons_container)

    def _create_basic_fields(self, parent):
        """Crear campos básicos con estilo profesional"""
        basic_frame = tk.LabelFrame(
            parent,
            text="Información Básica",
            font=("Segoe UI", 11, "bold"),
            bg="#ffffff",
            fg="#0B5394",
            bd=2,
            relief="groove"
        )
        basic_frame.pack(fill="both", expand=True, padx=2, pady=2, ipadx=6, ipady=6)

        # Configurar grid para expandir
        basic_frame.grid_columnconfigure(1, weight=1)

        tk.Label(basic_frame, text="# de muestra:", bg="#ffffff", font=("Segoe UI", 9, "bold"), fg="#0B5394").grid(row=0, column=0, sticky="w", padx=8, pady=4)
        self.parent.nombre_entry = tk.Entry(basic_frame, font=("Segoe UI", 9), relief="solid", bd=1)
        self.parent.nombre_entry.grid(row=0, column=1, padx=8, pady=4, sticky="ew")

        tk.Label(basic_frame, text="Descripción:", bg="#ffffff", font=("Segoe UI", 9, "bold"), fg="#0B5394").grid(row=1, column=0, sticky="nw", padx=8, pady=4)
        self.parent.descripcion_entry = tk.Text(basic_frame, font=("Segoe UI", 9), height=3, relief="solid", bd=1)
        self.parent.descripcion_entry.grid(row=1, column=1, padx=8, pady=4, sticky="ew")

        tk.Label(basic_frame, text="Fecha:", bg="#ffffff", font=("Segoe UI", 9, "bold"), fg="#0B5394").grid(row=2, column=0, sticky="w", padx=8, pady=4)
        self.parent.fecha_entry = tk.Entry(basic_frame, font=("Segoe UI", 9), state="readonly", bg="#f0f0f0", relief="solid", bd=1)
        fecha_actual = datetime.datetime.now().strftime("%Y-%m-%d")
        self.parent.fecha_entry.config(state="normal")
        self.parent.fecha_entry.insert(0, fecha_actual)
        self.parent.fecha_entry.config(state="readonly")
        self.parent.fecha_entry.grid(row=2, column=1, padx=8, pady=4, sticky="ew")

        tk.Label(basic_frame, text="Hora:", bg="#ffffff", font=("Segoe UI", 9, "bold"), fg="#0B5394").grid(row=3, column=0, sticky="w", padx=8, pady=4)
        self.parent.hora_entry = tk.Entry(basic_frame, font=("Segoe UI", 9), state="readonly", bg="#f0f0f0", relief="solid", bd=1)
        hora_actual = datetime.datetime.now().strftime("%H:%M:%S")
        self.parent.hora_entry.config(state="normal")
        self.parent.hora_entry.insert(0, hora_actual)
        self.parent.hora_entry.config(state="readonly")
        self.parent.hora_entry.grid(row=3, column=1, padx=8, pady=4, sticky="ew")

    def _create_nutrimental_fields(self, parent):
        """Crear campos nutricionales agrupados y visualmente atractivos"""
        nutri_frame = tk.LabelFrame(
            parent,
            text="Datos Nutricionales",
            font=("Segoe UI", 11, "bold"),
            bg="#ffffff",
            fg="#0B5394",
            bd=2,
            relief="groove"
        )
        nutri_frame.pack(fill="both", expand=True, padx=2, pady=2, ipadx=6, ipady=6)

        # Configurar grid para expandir
        nutri_frame.grid_columnconfigure(1, weight=1)

        self.parent.nutri_vars = {}
        nutri_fields = [
            ("humedad", "Humedad (%)"),
            ("cenizas", "Cenizas (%)"),
            ("proteina", "Proteína (%)"),
            ("grasa_total", "Grasa total (%)"),
            ("grasa_trans", "Grasas trans (mg/100g)"),
            ("fibra_dietetica", "Fibra dietética (%)"),
            ("azucares", "Azúcares totales (%)"),
            ("azucares_anadidos", "Azúcares añadidos (%)"),
            ("sodio", "Sodio (mg/100g)"),
            ("acidos_grasos_saturados", "Ácidos grasos saturados (%)"),
            ("porcion", "Tamaño de porción (g o mL)"),
            ("contenido_neto", "Contenido neto del envase (opcional)")
        ]

        for i, (key, label) in enumerate(nutri_fields):
            tk.Label(nutri_frame, text=label, bg="#ffffff", font=("Segoe UI", 9, "bold"), fg="#0B5394").grid(row=i, column=0, sticky="w", padx=8, pady=3)
            entry = tk.Entry(nutri_frame, font=("Segoe UI", 9), relief="solid", bd=1)
            entry.grid(row=i, column=1, padx=8, pady=3, sticky="ew")
            self.parent.nutri_vars[key] = entry

    def _create_results_area(self, parent):
        """Área de resultados con diseño visual atractivo"""
        self.parent.resultados_frame = tk.LabelFrame(
            parent,
            text="Resultados",
            font=("Segoe UI", 11, "bold"),
            bg="#ffffff",
            fg="#0B5394",
            bd=2,
            relief="groove"
        )
        self.parent.resultados_frame.pack(fill="both", expand=True, padx=2, pady=2, ipadx=6, ipady=6)

        self.parent.resultados_text = tk.Text(
            self.parent.resultados_frame,
            font=("Courier New", 9),
            bg="#f8f9fa",
            state="disabled",
            relief="solid",
            bd=1,
            wrap="word"
        )
        resultados_scroll = tk.Scrollbar(self.parent.resultados_frame, orient="vertical", command=self.parent.resultados_text.yview)
        self.parent.resultados_text.configure(yscrollcommand=resultados_scroll.set)

        self.parent.resultados_text.pack(side="left", fill="both", expand=True, padx=8, pady=8)
        resultados_scroll.pack(side="right", fill="y", pady=8)

    def _create_buttons(self, parent):
        """Botones con diseño moderno y agrupados"""
        calc_btn = tk.Button(
            parent,
            text="Calcular Tabla Nutrimental",
            command=self.calcular_tabla_nutrimental,
            bg="#0B5394",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            relief="flat",
            padx=15,
            pady=8,
            cursor="hand2",
            activebackground="#073763",
            activeforeground="white"
        )
        calc_btn.pack(side="left", padx=(0, 10))

        formato_btn = tk.Button(
            parent,
            text="Exportar en formato oficial",
            command=self.exportar_a_formato_predefinido,
            bg="#ffc107",
            fg="black",
            font=("Segoe UI", 11, "bold"),
            relief="flat",
            padx=15,
            pady=8,
            cursor="hand2",
            activebackground="#e0a800",
            activeforeground="black"
        )
        formato_btn.pack(side="left", padx=10)

        guardar_bd_btn = tk.Button(
            parent,
            text="Guardar en Base de Datos",
            command=self.guardar_solo_bd,
            bg="#007bff",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            relief="flat",
            padx=15,
            pady=8,
            cursor="hand2",
            activebackground="#0056b3",
            activeforeground="white"
        )
        guardar_bd_btn.pack(side="left", padx=10)

    def guardar_solo_bd(self):
        """Guardar solo en la base de datos, generando y guardando el PDF"""
        if not hasattr(self.parent, "ultimo_calculo"):
            messagebox.showwarning("Advertencia", "Primero debe calcular la tabla nutrimental")
            return

        try:
            # Verificar datos básicos
            nombre_actual = self.parent.nombre_entry.get().strip()
            descripcion_actual = self.parent.descripcion_entry.get("1.0", "end-1c").strip()
            if not nombre_actual:
                messagebox.showerror("Error", "El campo '# de muestra' es obligatorio para guardar")
                self.parent.nombre_entry.focus()
                return

            # Verificar usuario válido
            usuario_id = self.parent.get_usuario_id()
            if usuario_id is None:
                messagebox.showwarning("Advertencia", "No se pudo guardar en la base de datos: Usuario no válido")
                return

            # Cargar plantilla Excel
            ruta_plantilla = "formato.xlsx"
            wb = load_workbook(ruta_plantilla)
            ws = wb.active

            resultados = self.parent.ultimo_calculo["resultados"]
            entrada = self.parent.ultimo_calculo["datos_entrada"]
            datos_basicos = self.parent.ultimo_calculo["datos_basicos"]

            # Llenar datos en la plantilla Excel
            escribir_celda_segura(ws, "E17", f"{entrada.get('porcion', '')} g")
            escribir_celda_segura(ws, "D18", round(resultados.get("porciones_envase", 0), 1) if resultados.get("porciones_envase") else "")
            escribir_celda_segura(ws, "E18", "Porciones por envase:")
            escribir_celda_segura(ws, "F19", f"{resultados.get('por_envase', {}).get('energia_kcal', '')} kcal")
            escribir_celda_segura(ws, "F12", f"{entrada.get('contenido_neto', '')} g")
            escribir_celda_segura(ws, "C8", f"{nombre_actual} - {descripcion_actual}")

            # Por 100g/mL
            ws["F21"] = resultados["por_100g"].get("energia_kcal", "")
            ws["G21"] = "kcal"
            ws["F22"] = resultados["por_100g"].get("proteina", "")
            ws["G22"] = "g"
            ws["F23"] = resultados["por_100g"].get("grasa_total", "")
            ws["G23"] = "g"
            ws["F24"] = resultados["por_100g"].get("grasa_saturada", "")
            ws["G24"] = "g"
            ws["F25"] = resultados["por_100g"].get("grasa_trans", "")
            ws["G25"] = "mg"
            ws["F26"] = resultados["por_100g"].get("carbohidratos_disponibles", "")
            ws["G26"] = "g"
            ws["F27"] = resultados["por_100g"].get("azucares", "")
            ws["G27"] = "g"
            ws["F28"] = resultados["por_100g"].get("azucares_anadidos", "")
            ws["G28"] = "g"
            ws["F29"] = resultados["por_100g"].get("fibra_dietetica", "")
            ws["G29"] = "g"
            ws["F30"] = resultados["por_100g"].get("sodio", "")
            ws["G30"] = "mg"

            # Por porción
            ws["H21"] = resultados["por_porcion"].get("energia_kcal", "")
            ws["I21"] = "kcal"
            ws["H22"] = resultados["por_porcion"].get("proteina", "")
            ws["I22"] = "g"
            ws["H23"] = resultados["por_porcion"].get("grasa_total", "")
            ws["I23"] = "g"
            ws["H24"] = resultados["por_porcion"].get("grasa_saturada", "")
            ws["I24"] = "g"
            ws["H25"] = resultados["por_porcion"].get("grasa_trans", "")
            ws["I25"] = "mg"
            ws["H26"] = resultados["por_porcion"].get("carbohidratos_disponibles", "")
            ws["I26"] = "g"
            ws["H27"] = resultados["por_porcion"].get("azucares", "")
            ws["I27"] = "g"
            ws["H28"] = resultados["por_porcion"].get("azucares_anadidos", "")
            ws["I28"] = "g"
            ws["H29"] = resultados["por_porcion"].get("fibra_dietetica", "")
            ws["I29"] = "g"
            ws["H30"] = resultados["por_porcion"].get("sodio", "")
            ws["I30"] = "mg"

            # Crear archivos temporales
            tmp_excel_path = None
            tmp_pdf_path = None

            try:
                # Guardar Excel temporal
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_excel:
                    wb.save(tmp_excel.name)
                    tmp_excel_path = tmp_excel.name

                # Convertir a PDF
                import win32com.client
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                wb_pdf = excel.Workbooks.Open(tmp_excel_path)
                
                # Generar nombre para el PDF
                fecha = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                nombre_archivo_pdf = f"nutrimental_{nombre_actual}_{fecha}.pdf"
                tmp_pdf_path = os.path.join(tempfile.gettempdir(), nombre_archivo_pdf)
                
                # Exportar a PDF y cerrar Excel
                wb_pdf.ExportAsFixedFormat(0, tmp_pdf_path)
                wb_pdf.Close(False)
                excel.Quit()

                # Verificar que el PDF se creó correctamente
                if not os.path.exists(tmp_pdf_path):
                    raise Exception("No se pudo generar el archivo PDF")

                # Leer el PDF para guardarlo en la base de datos
                with open(tmp_pdf_path, "rb") as f:
                    archivo_binario = f.read()

                # Crear descripción para la base de datos
                fecha_actual = datetime.datetime.now()
                descripcion_completa = f"{descripcion_actual}\n\n"
                descripcion_completa += "TABLA NUTRIMENTAL OFICIAL (PDF):\n"
                descripcion_completa += f"- Proteína: {resultados['por_100g']['proteina']}g/100g\n"
                descripcion_completa += f"- Grasa total: {resultados['por_100g']['grasa_total']}g/100g\n"
                descripcion_completa += f"- Carbohidratos disponibles: {resultados['por_100g']['carbohidratos_disponibles']}g/100g\n"
                descripcion_completa += f"- Energía: {resultados['por_100g']['energia_kcal']} kcal/100g\n"
                descripcion_completa += f"- Tamaño de porción: {entrada['porcion']}g\n"
                descripcion_completa += f"- Contenido neto: {entrada.get('contenido_neto', 'No especificado')}\n"
                descripcion_completa += f"Archivo PDF formato oficial generado el {fecha_actual.strftime('%d/%m/%Y %H:%M:%S')}"

                # Guardar en la base de datos
                agregar_historial(
                    nombre_actual,
                    descripcion_completa,
                    fecha_actual.strftime("%Y-%m-%d"),
                    fecha_actual.strftime("%H:%M:%S"),
                    usuario_id,
                    archivo_binario
                )

                messagebox.showinfo(
                    "Guardado en Base de Datos",
                    f"✅ PDF nutrimental guardado correctamente en la base de datos.\n\n"
                    f"👤 Usuario: {self.parent.username}\n"
                    f"📝 # de muestra: {nombre_actual}\n"
                    f"📊 Formato: Tabla Nutrimental Oficial\n\n"
                    f"⚠️ No se generó archivo local."
                )

            except ImportError:
                messagebox.showwarning("PDF no generado", "No se pudo importar win32com. Instala con:\npip install pywin32")
            except Exception as e:
                messagebox.showerror("Error", f"Error al guardar PDF en base de datos: {str(e)}")
            finally:
                # Limpiar archivos temporales
                if tmp_excel_path and os.path.exists(tmp_excel_path):
                    try:
                        os.remove(tmp_excel_path)
                    except:
                        pass
                if tmp_pdf_path and os.path.exists(tmp_pdf_path):
                    try:
                        os.remove(tmp_pdf_path)
                    except:
                        pass

        except FileNotFoundError:
            messagebox.showerror("Error", "No se encontró la plantilla formato.xlsx. Debe estar en la carpeta principal del programa.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo completar la exportación:\n{str(e)}")

    def calcular_tabla_nutrimental(self):
        """Calcular tabla nutrimental"""
        try:
            # Obtener valores de los campos
            data = {}
            for key, entry in self.parent.nutri_vars.items():
                value = entry.get().strip()
                if value and key != "contenido_neto":
                    data[key] = float(value)
                elif value and key == "contenido_neto":
                    data[key] = float(value) if value else None

            # Campos requeridos
            required_fields = [
                "humedad", "cenizas", "proteina", "grasa_total", "grasa_trans", 
                "fibra_dietetica", "azucares", "azucares_anadidos", "sodio", 
                "acidos_grasos_saturados", "porcion"
            ]
            for field in required_fields:
                if field not in data:
                    field_name = field.replace('_', ' ').title()
                    messagebox.showerror("Error", f"El campo '{field_name}' es requerido")
                    return

            # Cálculos
            resultados = self._calcular_nutrimental(data)
            
            # Mostrar resultados
            self._mostrar_resultados(resultados)
            
            # Guardar para exportar
            self.parent.ultimo_calculo = {
                "datos_basicos": {
                    "nombre": self.parent.nombre_entry.get(),
                    "descripcion": self.parent.descripcion_entry.get("1.0", "end-1c"),
                    "fecha": self.parent.fecha_entry.get(),
                    "hora": self.parent.hora_entry.get()
                },
                "datos_entrada": data,
                "resultados": resultados
            }

        except ValueError:
            messagebox.showerror("Error", "Por favor ingrese valores numéricos válidos")
        except Exception as e:
            messagebox.showerror("Error", f"Error en el cálculo: {str(e)}")

    def _aplicar_redondeo_nutrientes(self, valor):
        """Aplicar reglas de redondeo para nutrientes en gramos según NOM-051"""
        if valor < 0.5:
            return 0
        elif valor <= 5:
            return round(valor * 2) / 2  # Redondear a 0.5g más cercano
        else:
            return round(valor)  # Redondear a gramo entero

    def _aplicar_redondeo_energia(self, valor):
        """Aplicar reglas de redondeo para energía según NOM-051"""
        if valor < 5:
            return 0
        elif valor <= 50:
            return round(valor / 5) * 5  # Redondear a múltiplos de 5
        else:
            return round(valor / 10) * 10  # Redondear a múltiplos de 10

    def _aplicar_regla_redondeo_sodio(self, valor):
        """
        Redondeo oficial para sodio y grasas trans (mg) según NOM-051:
        - < 5 mg: reportar 0
        - 5 a < 140 mg: redondear al múltiplo de 5 mg más cercano
        - >= 140 mg: redondear al múltiplo de 10 mg más cercano
        """
        if valor < 5:
            return 0
        elif valor < 140:
            return int(round(valor / 5) * 5)
        else:
            return int(round(valor / 10) * 10)

    def _calcular_nutrimental(self, data):
        """Cálculo nutrimental conforme NOM-051"""
        
        # 1. Aplicar reglas de redondeo para valores base
        proteina_100g = self._aplicar_redondeo_nutrientes(data["proteina"])
        grasa_total_100g = self._aplicar_redondeo_nutrientes(data["grasa_total"])
        fibra_dietetica_100g = self._aplicar_redondeo_nutrientes(data["fibra_dietetica"])
        azucares_100g = self._aplicar_redondeo_nutrientes(data["azucares"])
        azucares_anadidos_100g = self._aplicar_redondeo_nutrientes(data["azucares_anadidos"])
        
        # 2. Grasa saturada 
        grasa_saturada_100g = self._aplicar_redondeo_nutrientes(
            grasa_total_100g * (data["acidos_grasos_saturados"] / 100)
        )
        
        # 3. Hidratos de carbono
        hidratos_carbono_totales_100g = max(0, 100 - (
            data["humedad"] + data["cenizas"] + proteina_100g + grasa_total_100g
        ))
        carbohidratos_disponibles_100g = max(0, hidratos_carbono_totales_100g - fibra_dietetica_100g)
        
        # 4. Sodio y grasas trans 
        sodio_100g = self._aplicar_regla_redondeo_sodio(data["sodio"])
        grasa_trans_100g = self._aplicar_regla_redondeo_sodio(data["grasa_trans"])
        
        # CORRECTO: usar el valor original para la regla de 3 y luego redondear
        grasa_trans_porcion = self._aplicar_regla_redondeo_sodio(data["grasa_trans"] * data["porcion"] / 100)

        # 5. Energía por 100g (usar valores ya redondeados)
        energia_kcal_100g = int(round((proteina_100g + carbohidratos_disponibles_100g) * 4 + grasa_total_100g * 9))
        energia_kj_100g = int(round((proteina_100g + carbohidratos_disponibles_100g) * 17 + grasa_total_100g * 37))

        # 6. Por porción (regla de 3, usando valores ya redondeados de 100g, luego redondear a entero)
        porcion = data["porcion"]
        proteina_porcion = int(round(proteina_100g * porcion / 100))
        grasa_total_porcion = int(round(grasa_total_100g * porcion / 100))
        grasa_saturada_porcion = int(round(grasa_saturada_100g * porcion / 100))
        fibra_dietetica_porcion = int(round(fibra_dietetica_100g * porcion / 100))
        
        # Redondeo para azúcares y azúcares añadidos por porción
        azucares_porcion = int(round(data["azucares"] * porcion / 100))
        if azucares_porcion < 1:
            azucares_porcion = 0

        azucares_anadidos_porcion = int(round(data["azucares_anadidos"] * porcion / 100))
        if azucares_anadidos_porcion < 1:
            azucares_anadidos_porcion = 0
            
        carbohidratos_disponibles_porcion = int(round(carbohidratos_disponibles_100g * porcion / 100))
        sodio_porcion = self._aplicar_regla_redondeo_sodio(sodio_100g * porcion / 100)
        grasa_trans_porcion = self._aplicar_regla_redondeo_sodio(grasa_trans_100g * porcion / 100)

        # Energía por porción (usar valores ya redondeados de porción)
        energia_kcal_porcion = int(round((proteina_porcion + carbohidratos_disponibles_porcion) * 4 + grasa_total_porcion * 9))
        energia_kj_porcion = int(round((proteina_porcion + carbohidratos_disponibles_porcion) * 17 + grasa_total_porcion * 37))

        # 7. Porciones por envase (solo aquí se permiten decimales)
        porciones_envase = None
        if "contenido_neto" in data and data["contenido_neto"] and porcion:
            try:
                porciones_envase = data["contenido_neto"] / porcion
            except Exception:
                porciones_envase = None

        # 8. Por envase (energía)
        por_envase = None
        if "contenido_neto" in data and data["contenido_neto"]:
            factor_cn = data["contenido_neto"] / 100
            energia_kcal_envase = int(round(energia_kcal_100g * factor_cn))
            energia_kj_envase = int(round(energia_kj_100g * factor_cn))
            por_envase = {
                "energia_kcal": energia_kcal_envase,
                "energia_kj": energia_kj_envase
            }

        resultados = {
            "por_100g": {
                "proteina": proteina_100g,
                "grasa_total": grasa_total_100g,
                "grasa_saturada": grasa_saturada_100g,
                "grasa_trans": grasa_trans_100g,
                "carbohidratos_disponibles": carbohidratos_disponibles_100g,
                "azucares": azucares_100g,
                "azucares_anadidos": azucares_anadidos_100g,
                "fibra_dietetica": fibra_dietetica_100g,
                "sodio": sodio_100g,
                "energia_kcal": energia_kcal_100g,
                "energia_kj": energia_kj_100g
            },
            "por_porcion": {
                "proteina": proteina_porcion,
                "grasa_total": grasa_total_porcion,
                "grasa_saturada": grasa_saturada_porcion,
                "grasa_trans": grasa_trans_porcion,
                "carbohidratos_disponibles": carbohidratos_disponibles_porcion,
                "azucares": azucares_porcion,
                "azucares_anadidos": azucares_anadidos_porcion,
                "fibra_dietetica": fibra_dietetica_porcion,
                "sodio": sodio_porcion,
                "energia_kcal": energia_kcal_porcion,
                "energia_kj": energia_kj_porcion
            }
        }

        if porciones_envase is not None:
            resultados["porciones_envase"] = porciones_envase

        if por_envase is not None:
            resultados["por_envase"] = por_envase

        resultados["es_porcion_100g"] = porcion == 100.0

        return resultados

    def _mostrar_resultados(self, resultados):
        """Mostrar resultados según formato de tabla nutrimental mexicana"""
        self.parent.resultados_text.config(state="normal")
        self.parent.resultados_text.delete("1.0", "end")
        
        texto = "TABLA NUTRIMENTAL MEXICANA\n"
        texto += "=" * 50 + "\n\n"

        # Porciones por envase - CON DECIMALES
        if "porciones_envase" in resultados and resultados["porciones_envase"] is not None:
            porciones = resultados["porciones_envase"]
            if porciones == int(porciones):
                texto += f"Porciones por envase: {int(porciones)}\n\n"
            else:
                texto += f"Porciones por envase: {porciones:.1f}\n\n"

        # Por 100g/mL
        texto += "POR 100g/mL:\n"
        texto += "-" * 20 + "\n"
        
        # Orden según normativa mexicana
        orden_campos = [
            ("proteina", "Proteína", "g"),
            ("grasa_total", "Grasa Total", "g"), 
            ("grasa_saturada", "Grasa Saturada", "g"),
            ("grasa_trans", "Grasas trans", "mg"),
            ("carbohidratos_disponibles", "Carbohidratos Disponibles", "g"),
            ("azucares", "Azúcares", "g"),
            ("azucares_anadidos", "Azúcares añadidos", "g"),
            ("fibra_dietetica", "Fibra Dietética", "g"),
            ("sodio", "Sodio", "mg"),
            ("energia_kcal", "Contenido energético", "kcal"),
            ("energia_kj", "Contenido energético", "kJ")
        ]
        
        for key, nombre, unidad in orden_campos:
            if key in resultados["por_100g"]:
                value = resultados["por_100g"][key]
                texto += f"{nombre}: {value} {unidad}\n"
        texto += "\n"
        
        # Por porción - SOLO SI NO ES 100g/mL
        if not resultados.get("es_porcion_100g", False):
            texto += "POR PORCIÓN:\n"
            texto += "-" * 20 + "\n"
            for key, nombre, unidad in orden_campos:
                if key in resultados["por_porcion"]:
                    value = resultados["por_porcion"][key]
                    texto += f"{nombre}: {value} {unidad}\n"
            texto += "\n"
        
        # Por envase
        if "por_envase" in resultados:
            texto += "POR ENVASE:\n"
            texto += "-" * 20 + "\n"
            texto += f"Contenido energético: {resultados['por_envase']['energia_kcal']} kcal\n"
            texto += f"Contenido energético: {resultados['por_envase']['energia_kj']} kJ\n"

        self.parent.resultados_text.insert("1.0", texto)
        self.parent.resultados_text.config(state="disabled")

    def exportar_nutrimental_excel(self):
        """Exportar tabla nutrimental a Excel"""
        if not hasattr(self.parent, "ultimo_calculo"):
            messagebox.showwarning("Advertencia", "Primero debe calcular la tabla nutrimental")
            return
        
        try:
            # Obtener valores actuales
            nombre_actual = self.parent.nombre_entry.get().strip()
            descripcion_actual = self.parent.descripcion_entry.get("1.0", "end-1c").strip()
            
            if not nombre_actual:
                messagebox.showerror("Error", "El campo '# de muestra' es obligatorio para exportar")
                self.parent.nombre_entry.focus()
                return
            
            # Generar nombre de archivo
            fecha_actual = datetime.datetime.now()
            nombre_limpio = "".join(c for c in nombre_actual if c.isalnum() or c in (' ', '-', '_')).rstrip()
            nombre_limpio = nombre_limpio.replace(' ', '_')
            if not nombre_limpio:
                nombre_limpio = "Tabla_Nutrimental"
            
            timestamp = fecha_actual.strftime("%Y-%m-%d_%H%M%S")
            nombre_predeterminado = f"{nombre_limpio}_{timestamp}.xlsx"
            
            # Preparar datos para Excel
            datos_excel = []
            
            # Información básica
            datos_excel.append(["INFORMACIÓN BÁSICA", "", ""])
            datos_excel.append(["# de muestra", nombre_actual, ""])
            datos_excel.append(["Descripción", descripcion_actual, ""])
            datos_excel.append(["Fecha de análisis", fecha_actual.strftime("%Y-%m-%d"), ""])
            datos_excel.append(["Hora de análisis", fecha_actual.strftime("%H:%M:%S"), ""])
            datos_excel.append(["Usuario", self.parent.username, ""])
            datos_excel.append(["Fecha de exportación", fecha_actual.strftime("%Y-%m-%d %H:%M:%S"), ""])
            datos_excel.append(["", "", ""])
            
            # Datos de entrada
            datos_excel.append(["DATOS DE ENTRADA", "", ""])
            for key, value in self.parent.ultimo_calculo["datos_entrada"].items():
                nombre_campo = key.replace('_', ' ').title()
                if key == "sodio":
                    datos_excel.append([f"{nombre_campo} (mg/100g)", value, ""])
                elif key == "grasa_trans":
                    datos_excel.append([f"{nombre_campo} (mg/100g)", value, ""])
                elif key == "porcion":
                    datos_excel.append([f"{nombre_campo} (g)", value, ""])
                elif key == "contenido_neto":
                    datos_excel.append([f"{nombre_campo} (g/mL)", value, ""])
                else:
                    datos_excel.append([f"{nombre_campo} (%)", value, ""])
            datos_excel.append(["", "", ""])
            
            # Tabla nutrimental
            datos_excel.append(["TABLA NUTRIMENTAL MEXICANA", "Por 100g/mL", "Por Porción"])
            resultados = self.parent.ultimo_calculo["resultados"]

            # Porciones por envase
            if "porciones_envase" in resultados and resultados["porciones_envase"] is not None:
                datos_excel.append(["Porciones por envase", resultados["porciones_envase"], ""])

            for key in resultados["por_100g"].keys():
                if key == "energia_kcal":
                    nombre = "Contenido energético (kcal)"
                elif key == "energia_kj":
                    nombre = "Contenido energético (kJ)"
                elif key == "sodio":
                    nombre = "Sodio (mg)"
                elif key == "grasa_trans":
                    nombre = "Grasas trans (mg)"  # Modificado
                elif key == "azucares_anadidos":
                    nombre = "Azúcares añadidos (g)"  # Modificado
                else:
                    nombre = f"{key.replace('_', ' ').title()} (g)"
                valor_100g = resultados["por_100g"][key]
                valor_porcion = resultados["por_porcion"][key]
                datos_excel.append([nombre, valor_100g, valor_porcion])
            
            # Por envase si existe
            if "por_envase" in resultados:
                datos_excel.append(["", "", ""])
                datos_excel.append(["POR ENVASE COMPLETO", "", ""])
                for key, value in resultados["por_envase"].items():
                    if key == "energia_kcal":
                        nombre = "Contenido energético total (kcal)"
                    elif key == "energia_kj":
                        nombre = "Contenido energético total (kJ)"
                    else:
                        nombre = f"{key.replace('_', ' ').title()}"
                    datos_excel.append([nombre, value, ""])
            
            # Crear DataFrame
            df = pd.DataFrame(datos_excel, columns=["Componente", "Valor 100g/mL", "Valor Porción"])
            
            # Ubicación de descarga
            try:
                desktop = os.path.join(os.path.expanduser("~"), "Desktop")
                if not os.path.exists(desktop):
                    desktop = os.path.expanduser("~")
            except:
                desktop = ""
            
            # Guardar archivo
            filename = filedialog.asksaveasfilename(
                initialfile=nombre_predeterminado,
                initialdir=desktop,
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Guardar tabla nutrimental"
            )
            
            if filename:
                # Crear archivo Excel
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name="Tabla Nutrimental")
                    
                    workbook = writer.book
                    worksheet = writer.sheets["Tabla Nutrimental"]
                    
                    # Ajustar ancho de columnas
                    worksheet.column_dimensions['A'].width = 30
                    worksheet.column_dimensions['B'].width = 15
                    worksheet.column_dimensions['C'].width = 15
                
                # Leer archivo para base de datos
                with open(filename, "rb") as f:
                    archivo_binario = f.read()
                
                # Guardar en base de datos
                usuario_id = self.parent.get_usuario_id()
                
                if usuario_id is None:
                    messagebox.showwarning("Advertencia", "No se pudo guardar en la base de datos: Usuario no válido")
                    return
                
                # Descripción completa
                descripcion_completa = f"{descripcion_actual}\n\n"
                descripcion_completa += "DATOS NUTRICIONALES CALCULADOS:\n"
                descripcion_completa += f"- Proteína: {resultados['por_100g']['proteina']}g/100g\n"
                descripcion_completa += f"- Grasa total: {resultados['por_100g']['grasa_total']}g/100g\n"
                descripcion_completa += f"- Carbohidratos disponibles: {resultados['por_100g']['carbohidratos_disponibles']}g/100g\n"
                descripcion_completa += f"- Energía: {resultados['por_100g']['energia_kcal']} kcal/100g\n"
                descripcion_completa += f"- Tamaño de porción analizada: {self.parent.ultimo_calculo['datos_entrada']['porcion']}g\n"
                descripcion_completa += f"Archivo Excel generado automáticamente: {os.path.basename(filename)}"
                
                agregar_historial(
                    nombre_actual,
                    descripcion_completa,
                    fecha_actual.strftime("%Y-%m-%d"),
                    fecha_actual.strftime("%H:%M:%S"),
                    usuario_id,
                    archivo_binario
                )
                
                messagebox.showinfo(
                    "Exportación Exitosa", 
                    f"✅ Tabla nutrimental exportada correctamente:\n\n"
                    f"📁 Archivo: {os.path.basename(filename)}\n"
                    f"📂 Ubicación: {filename}\n"
                    f"💾 Guardado en base de datos: ✓\n"
                    f"👤 Usuario: {self.parent.username}\n"
                    f"📝 # de muestra: {nombre_actual}\n\n"
                )
            else:
                messagebox.showinfo("Cancelado", "Exportación cancelada por el usuario.")
        
        except ImportError:
            messagebox.showerror("Error", "No se pudo importar pandas. Asegúrate de que esté instalado:\npip install pandas openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {str(e)}")

    def exportar_a_formato_predefinido(self):
        import sys
        if not hasattr(self.parent, "ultimo_calculo"):
            messagebox.showwarning("Advertencia", "Primero debe calcular la tabla nutrimental")
            return

        try:
            ruta_plantilla = "formato.xlsx"
            wb = load_workbook(ruta_plantilla)
            ws = wb.active

            resultados = self.parent.ultimo_calculo["resultados"]
            entrada = self.parent.ultimo_calculo["datos_entrada"]
            datos_basicos = self.parent.ultimo_calculo["datos_basicos"]

            # === Valores generales ===
            escribir_celda_segura(ws, "E17", f"{entrada.get('porcion', '')} g")
            escribir_celda_segura(ws, "D18", round(resultados.get("porciones_envase", 0), 1) if resultados.get("porciones_envase") else "")
            escribir_celda_segura(ws, "E18", "Porciones por envase:")
            escribir_celda_segura(ws, "F19", f"{resultados.get('por_envase', {}).get('energia_kcal', '')} kcal")
            escribir_celda_segura(ws, "F12", f"{entrada.get('contenido_neto', '')} g")
            nombre_muestra = datos_basicos.get("nombre", "")
            descripcion = datos_basicos.get("descripcion", "")
            escribir_celda_segura(ws, "C8", f"{nombre_muestra} - {descripcion}")

            # === Por 100 mL ===
            ws["F21"] = resultados["por_100g"].get("energia_kcal", "")
            ws["G21"] = "kcal"
            ws["F22"] = resultados["por_100g"].get("proteina", "")
            ws["G22"] = "g"
            ws["F23"] = resultados["por_100g"].get("grasa_total", "")
            ws["G23"] = "g"
            ws["F24"] = resultados["por_100g"].get("grasa_saturada", "")
            ws["G24"] = "g"
            ws["F25"] = resultados["por_100g"].get("grasa_trans", "")
            ws["G25"] = "mg"
            ws["F26"] = resultados["por_100g"].get("carbohidratos_disponibles", "")
            ws["G26"] = "g"
            ws["F27"] = resultados["por_100g"].get("azucares", "")
            ws["G27"] = "g"
            ws["F28"] = resultados["por_100g"].get("azucares_anadidos", "")
            ws["G28"] = "g"
            ws["F29"] = resultados["por_100g"].get("fibra_dietetica", "")
            ws["G29"] = "g"
            ws["F30"] = resultados["por_100g"].get("sodio", "")
            ws["G30"] = "mg"

            # === Por porción ===
            ws["H21"] = resultados["por_porcion"].get("energia_kcal", "")
            ws["I21"] = "kcal"
            ws["H22"] = resultados["por_porcion"].get("proteina", "")
            ws["I22"] = "g"
            ws["H23"] = resultados["por_porcion"].get("grasa_total", "")
            ws["I23"] = "g"
            ws["H24"] = resultados["por_porcion"].get("grasa_saturada", "")
            ws["I24"] = "g"
            ws["H25"] = resultados["por_porcion"].get("grasa_trans", "")
            ws["I25"] = "mg"
            ws["H26"] = resultados["por_porcion"].get("carbohidratos_disponibles", "")
            ws["I26"] = "g"
            ws["H27"] = resultados["por_porcion"].get("azucares", "")
            ws["I27"] = "g"
            ws["H28"] = resultados["por_porcion"].get("azucares_anadidos", "")
            ws["I28"] = "g"
            ws["H29"] = resultados["por_porcion"].get("fibra_dietetica", "")
            ws["I29"] = "g"
            ws["H30"] = resultados["por_porcion"].get("sodio", "")
            ws["I30"] = "mg"

            # === Guardar archivo Excel temporal ===
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_excel:
                wb.save(tmp_excel.name)
                tmp_excel_path = tmp_excel.name

            # === Convertir a PDF usando Excel (solo Windows con Office) ===
            try:
                import win32com.client
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                wb_pdf = excel.Workbooks.Open(tmp_excel_path)
                fecha = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                nombre_archivo_pdf = f"nutrimental_{nombre_muestra}_{fecha}.pdf"
                ruta_guardado_pdf = filedialog.asksaveasfilename(
                    defaultextension=".pdf",
                    initialfile=nombre_archivo_pdf,
                    filetypes=[("Archivos PDF", "*.pdf")]
                )
                if not ruta_guardado_pdf:
                    wb_pdf.Close(False)
                    excel.Quit()
                    os.remove(tmp_excel_path)
                    messagebox.showinfo("Cancelado", "Exportación cancelada por el usuario.")
                    return
                wb_pdf.ExportAsFixedFormat(0, ruta_guardado_pdf)
                wb_pdf.Close(False)
                excel.Quit()
                os.remove(tmp_excel_path)
                messagebox.showinfo("Éxito", f"Archivo PDF exportado correctamente:\n\n{ruta_guardado_pdf}")
            except ImportError:
                os.remove(tmp_excel_path)
                messagebox.showwarning("PDF no generado", "No se pudo importar win32com. Instala con:\npip install pywin32")
            except Exception as e:
                os.remove(tmp_excel_path)
                messagebox.showerror("Error PDF", f"No se pudo convertir a PDF:\n{str(e)}")

        except FileNotFoundError:
            messagebox.showerror("Error", "No se encontró la plantilla formato.xlsx. Debe estar en la carpeta principal del programa.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo completar la exportación:\n{str(e)}")

    def guardar_solo_bd(self):
        """Guardar solo en la base de datos, generando y guardando el PDF"""
        if not hasattr(self.parent, "ultimo_calculo"):
            messagebox.showwarning("Advertencia", "Primero debe calcular la tabla nutrimental")
            return

        try:
            # Verificar datos básicos
            nombre_actual = self.parent.nombre_entry.get().strip()
            descripcion_actual = self.parent.descripcion_entry.get("1.0", "end-1c").strip()
            if not nombre_actual:
                messagebox.showerror("Error", "El campo '# de muestra' es obligatorio para guardar")
                self.parent.nombre_entry.focus()
                return

            # Verificar usuario válido
            usuario_id = self.parent.get_usuario_id()
            if usuario_id is None:
                messagebox.showwarning("Advertencia", "No se pudo guardar en la base de datos: Usuario no válido")
                return

            # Cargar plantilla Excel
            ruta_plantilla = "formato.xlsx"
            wb = load_workbook(ruta_plantilla)
            ws = wb.active

            resultados = self.parent.ultimo_calculo["resultados"]
            entrada = self.parent.ultimo_calculo["datos_entrada"]
            datos_basicos = self.parent.ultimo_calculo["datos_basicos"]

            # Llenar datos en la plantilla Excel
            escribir_celda_segura(ws, "E17", f"{entrada.get('porcion', '')} g")
            escribir_celda_segura(ws, "D18", round(resultados.get("porciones_envase", 0), 1) if resultados.get("porciones_envase") else "")
            escribir_celda_segura(ws, "E18", "Porciones por envase:")
            escribir_celda_segura(ws, "F19", f"{resultados.get('por_envase', {}).get('energia_kcal', '')} kcal")
            escribir_celda_segura(ws, "F12", f"{entrada.get('contenido_neto', '')} g")
            escribir_celda_segura(ws, "C8", f"{nombre_actual} - {descripcion_actual}")

            # Por 100g/mL
            ws["F21"] = resultados["por_100g"].get("energia_kcal", "")
            ws["G21"] = "kcal"
            ws["F22"] = resultados["por_100g"].get("proteina", "")
            ws["G22"] = "g"
            ws["F23"] = resultados["por_100g"].get("grasa_total", "")
            ws["G23"] = "g"
            ws["F24"] = resultados["por_100g"].get("grasa_saturada", "")
            ws["G24"] = "g"
            ws["F25"] = resultados["por_100g"].get("grasa_trans", "")
            ws["G25"] = "mg"
            ws["F26"] = resultados["por_100g"].get("carbohidratos_disponibles", "")
            ws["G26"] = "g"
            ws["F27"] = resultados["por_100g"].get("azucares", "")
            ws["G27"] = "g"
            ws["F28"] = resultados["por_100g"].get("azucares_anadidos", "")
            ws["G28"] = "g"
            ws["F29"] = resultados["por_100g"].get("fibra_dietetica", "")
            ws["G29"] = "g"
            ws["F30"] = resultados["por_100g"].get("sodio", "")
            ws["G30"] = "mg"

            # Por porción
            ws["H21"] = resultados["por_porcion"].get("energia_kcal", "")
            ws["I21"] = "kcal"
            ws["H22"] = resultados["por_porcion"].get("proteina", "")
            ws["I22"] = "g"
            ws["H23"] = resultados["por_porcion"].get("grasa_total", "")
            ws["I23"] = "g"
            ws["H24"] = resultados["por_porcion"].get("grasa_saturada", "")
            ws["I24"] = "g"
            ws["H25"] = resultados["por_porcion"].get("grasa_trans", "")
            ws["I25"] = "mg"
            ws["H26"] = resultados["por_porcion"].get("carbohidratos_disponibles", "")
            ws["I26"] = "g"
            ws["H27"] = resultados["por_porcion"].get("azucares", "")
            ws["I27"] = "g"
            ws["H28"] = resultados["por_porcion"].get("azucares_anadidos", "")
            ws["I28"] = "g"
            ws["H29"] = resultados["por_porcion"].get("fibra_dietetica", "")
            ws["I29"] = "g"
            ws["H30"] = resultados["por_porcion"].get("sodio", "")
            ws["I30"] = "mg"

            # Crear archivos temporales
            tmp_excel_path = None
            tmp_pdf_path = None

            try:
                # Guardar Excel temporal
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_excel:
                    wb.save(tmp_excel.name)
                    tmp_excel_path = tmp_excel.name

                # Convertir a PDF
                import win32com.client
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                wb_pdf = excel.Workbooks.Open(tmp_excel_path)
                
                # Generar nombre para el PDF
                fecha = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                nombre_archivo_pdf = f"nutrimental_{nombre_actual}_{fecha}.pdf"
                tmp_pdf_path = os.path.join(tempfile.gettempdir(), nombre_archivo_pdf)
                
                # Exportar a PDF y cerrar Excel
                wb_pdf.ExportAsFixedFormat(0, tmp_pdf_path)
                wb_pdf.Close(False)
                excel.Quit()

                # Verificar que el PDF se creó correctamente
                if not os.path.exists(tmp_pdf_path):
                    raise Exception("No se pudo generar el archivo PDF")

                # Leer el PDF para guardarlo en la base de datos
                with open(tmp_pdf_path, "rb") as f:
                    archivo_binario = f.read()

                # Crear descripción para la base de datos
                fecha_actual = datetime.datetime.now()
                descripcion_completa = f"{descripcion_actual}\n\n"
                descripcion_completa += "TABLA NUTRIMENTAL OFICIAL (PDF):\n"
                descripcion_completa += f"- Proteína: {resultados['por_100g']['proteina']}g/100g\n"
                descripcion_completa += f"- Grasa total: {resultados['por_100g']['grasa_total']}g/100g\n"
                descripcion_completa += f"- Carbohidratos disponibles: {resultados['por_100g']['carbohidratos_disponibles']}g/100g\n"
                descripcion_completa += f"- Energía: {resultados['por_100g']['energia_kcal']} kcal/100g\n"
                descripcion_completa += f"- Tamaño de porción: {entrada['porcion']}g\n"
                descripcion_completa += f"- Contenido neto: {entrada.get('contenido_neto', 'No especificado')}\n"
                descripcion_completa += f"Archivo PDF formato oficial generado el {fecha_actual.strftime('%d/%m/%Y %H:%M:%S')}"

                # Guardar en la base de datos
                agregar_historial(
                    nombre_actual,
                    descripcion_completa,
                    fecha_actual.strftime("%Y-%m-%d"),
                    fecha_actual.strftime("%H:%M:%S"),
                    usuario_id,
                    archivo_binario
                )

                messagebox.showinfo(
                    "Guardado en Base de Datos",
                    f"✅ PDF nutrimental guardado correctamente en la base de datos.\n\n"
                    f"👤 Usuario: {self.parent.username}\n"
                    f"📝 # de muestra: {nombre_actual}\n"
                    f"📊 Formato: Tabla Nutrimental Oficial\n\n"
                    f"⚠️ No se generó archivo local."
                )

            except ImportError:
                messagebox.showwarning("PDF no generado", "No se pudo importar win32com. Instala con:\npip install pywin32")
            except Exception as e:
                messagebox.showerror("Error", f"Error al guardar PDF en base de datos: {str(e)}")
            finally:
                # Limpiar archivos temporales
                if tmp_excel_path and os.path.exists(tmp_excel_path):
                    try:
                        os.remove(tmp_excel_path)
                    except:
                        pass
                if tmp_pdf_path and os.path.exists(tmp_pdf_path):
                    try:
                        os.remove(tmp_pdf_path)
                    except:
                        pass

        except FileNotFoundError:
            messagebox.showerror("Error", "No se encontró la plantilla formato.xlsx. Debe estar en la carpeta principal del programa.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo completar la exportación:\n{str(e)}")

    def _generar_pdf_desde_excel(self, wb, nombre_archivo, guardar_dialogo=True):
        """Función centralizada para generar PDF desde Excel"""
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_excel:
                wb.save(tmp_excel.name)
                tmp_excel_path = tmp_excel.name
                
            import win32com.client
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            wb_pdf = excel.Workbooks.Open(tmp_excel_path)
            
            if guardar_dialogo:
                ruta_guardado_pdf = filedialog.asksaveasfilename(
                    defaultextension=".pdf",
                    initialfile=nombre_archivo,
                    filetypes=[("Archivos PDF", "*.pdf")]
                )
                if not ruta_guardado_pdf:
                    return None, "Cancelado"
            else:
                ruta_guardado_pdf = os.path.join(tempfile.gettempdir(), nombre_archivo)
                
            wb_pdf.ExportAsFixedFormat(0, ruta_guardado_pdf)
            wb_pdf.Close(False)
            excel.Quit()
            
            with open(ruta_guardado_pdf, "rb") as f:
                archivo_binario = f.read()
                
            return archivo_binario, ruta_guardado_pdf
        finally:
            if 'tmp_excel_path' in locals() and os.path.exists(tmp_excel_path):
                os.remove(tmp_excel_path)